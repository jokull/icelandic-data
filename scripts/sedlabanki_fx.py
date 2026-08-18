"""
Seðlabanki FX intervention — CBI net purchases/sales, FX market turnover,
FX reserves, and the ISK/EUR mid rate, monthly since 2019.

Sources (all public, no auth):
  1. XML time series (direct HTTP — NOT the fr.sedlabanki.is SDMX proxy):
     - GroupID 8 "Velta á gjaldeyrismarkaði" (FX market, DAILY since Jan 2009):
         * TS 282  Heildarvelta, M.eur          (market turnover)
         * TS 284  Heildarvelta, M.kr.          (market turnover)
         * TS 285  Gjaldeyrissala SÍ í ISK      (CBI *sales* of FX, M.kr.)
         * TS 287  Gjaldeyriskaup SÍ í ISK      (CBI *purchases* of FX, M.kr.)
     - TS 4064  Evra, skráð miðgengi (EUR mid rate, DAILY)
  2. CBI balance sheet Excel ("Sedlabanki" sheet, MONTHLY since 1994-01) —
     the "Gjaldeyrisforði (erlendar eignir, a-f)" memo line is FX reserves.
     Served from sedlabanki.is/library through the gagnabanki.is/api/download
     proxy (same path the existing balance-sheets/newcredit fetches use).

The daily series are summed per calendar month. Reserves and the ISK/EUR mid
rate are month-end. `net_purchases_mkr` is positive when the CBI bought FX
(kaup) and negative when it sold (sala) — during COVID 2020-21 the CBI was a
net SELLER of FX (defending the króna), flipping to a net buyer from mid-2025.

Usage:
    uv run python scripts/sedlabanki_fx.py list        # show series found
    uv run python scripts/sedlabanki_fx.py fetch       # raw + processed CSV
"""

from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

import httpx
import polars as pl
from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

XML_BASE = "https://sedlabanki.is/xmltimeseries/Default.aspx"
FX_GROUP = 8          # Velta á gjaldeyrismarkaði
EUR_MID_TS = 4064     # Evra, skráð miðgengi
CB_BALANCE_LIBRARY = (
    "https://sedlabanki.is/library?itemid=c0126d81-fd88-42bd-aee3-449e09b9089f"
)
PROXY = "https://gagnabanki.is/api/download"
START_DATE = "2019-01-01"  # covers COVID (2020-21) + the 2025 carry-trade era

RAW_DIR = Path(__file__).parent.parent / "data" / "raw" / "sedlabanki"
PROCESSED_DIR = Path(__file__).parent.parent / "data" / "processed"

OUT_CSV = PROCESSED_DIR / "sedlabanki_fx_intervention.csv"

# group 8 series this pipeline depends on -> English label
FX_SERIES = {
    282: "FX market turnover, M.eur",
    284: "FX market turnover, M.kr",
    285: "CBI FX sales (gjaldeyrissala), M.kr",
    287: "CBI FX purchases (gjaldeyriskaup), M.kr",
}

# "Sedlabanki" sheet row numbers in the balance-sheet workbook (stable layout)
_CB_ROW_DATES = 9
_CB_ROW_RESERVES = 60  # Liðir til skýringar: Gjaldeyrisforði (erlendar eignir, a-f)


def _client() -> httpx.Client:
    return httpx.Client(timeout=60.0, follow_redirects=True)


def fetch_fx_market(client: httpx.Client) -> Path:
    """Daily FX market turnover + CBI purchases/sales (GroupID 8) as raw CSV."""
    params = {"DagsFra": START_DATE, "GroupID": FX_GROUP, "Type": "csv"}
    r = client.get(XML_BASE, params=params)
    r.raise_for_status()
    out = RAW_DIR / "fx_market_daily.csv"
    out.write_bytes(r.content)
    print(f"  {out.name}: {len(r.text.splitlines())} daily rows")
    return out


def fetch_eur_mid(client: httpx.Client) -> Path:
    """Daily EUR mid rate (TS 4064) as raw CSV."""
    params = {"DagsFra": START_DATE, "TimeSeriesID": EUR_MID_TS, "Type": "csv"}
    r = client.get(XML_BASE, params=params)
    r.raise_for_status()
    out = RAW_DIR / "eur_mid_daily.csv"
    out.write_bytes(r.content)
    print(f"  {out.name}: {len(r.text.splitlines())} daily rows")
    return out


def fetch_cb_balance_sheet(client: httpx.Client) -> Path:
    """CBI balance sheet workbook (monthly reserves) via the gagnabanki proxy."""
    r = client.post(PROXY, json={"url": CB_BALANCE_LIBRARY})
    r.raise_for_status()
    out = RAW_DIR / "cb_balance_sheets.xlsx"
    out.write_bytes(r.content)
    print(f"  {out.name}: {len(r.content)} bytes")
    return out


def read_fx_market(path: Path) -> pl.DataFrame:
    """Parse the semicolon CSV from xmltimeseries into a tidy long frame.

    Line format:
        group;group_name;series_id;;series_name;description;date;value
    """
    rows = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        parts = line.split(";")
        if len(parts) < 8:
            continue
        try:
            series_id = int(parts[2])
        except ValueError:
            continue
        if series_id not in FX_SERIES:
            continue
        try:
            value = float(parts[7])
        except ValueError:
            continue
        # dates look like "1/2/2019 12:00:00 AM"
        d = pl.Series([parts[6]]).str.to_date("%m/%d/%Y %I:%M:%S %p")[0]
        rows.append({"date": d, "series_id": series_id, "value_mkr": value})
    df = pl.DataFrame(rows, schema={
        "date": pl.Date,
        "series_id": pl.Int32,
        "value_mkr": pl.Float64,
    })
    if df.is_empty():
        raise RuntimeError(f"no known FX series in {path}")
    return df


def read_eur_mid(path: Path) -> pl.DataFrame:
    rows = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        parts = line.split(";")
        if len(parts) < 8:
            continue
        try:
            value = float(parts[7])
        except ValueError:
            continue
        d = pl.Series([parts[6]]).str.to_date("%m/%d/%Y %I:%M:%S %p")[0]
        rows.append({"date": d, "eur_mid": value})
    df = pl.DataFrame(rows, schema={"date": pl.Date, "eur_mid": pl.Float64})
    if df.is_empty():
        raise RuntimeError(f"no EUR mid data in {path}")
    return df


def read_reserves(path: Path) -> pl.DataFrame:
    """Month-end FX reserves (M.kr.) from the 'Sedlabanki' sheet."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Sedlabanki"]
    dates = [c for c in next(ws.iter_rows(min_row=_CB_ROW_DATES, max_row=_CB_ROW_DATES, values_only=True)) if c is not None][1:]
    reserve_row = next(ws.iter_rows(min_row=_CB_ROW_RESERVES, max_row=_CB_ROW_RESERVES, values_only=True))
    vals = reserve_row[2:2 + len(dates)]
    rows = []
    for d, v in zip(dates, vals):
        if v is None:
            continue
        if hasattr(d, "strftime"):
            dt = d.date() if hasattr(d, "date") else d
        else:
            dt = date.fromisoformat(str(d)[:10])
        rows.append({"date": dt, "reserves_mkr": float(v)})
    return pl.DataFrame(rows, schema={"date": pl.Date, "reserves_mkr": pl.Float64})


def build_monthly(fx_path: Path, eur_path: Path, cb_path: Path) -> pl.DataFrame:
    """Aggregate the three sources into one monthly wide frame."""
    fx = read_fx_market(fx_path)
    eur = read_eur_mid(eur_path)
    res = read_reserves(cb_path)

    fx = fx.with_columns((pl.col("date").dt.truncate("1mo")).alias("month"))
    eur = eur.with_columns((pl.col("date").dt.truncate("1mo")).alias("month"))
    res = res.with_columns((pl.col("date").dt.truncate("1mo")).alias("month"))

    monthly = (
        fx.pivot(
            values="value_mkr",
            index="month",
            on="series_id",
            aggregate_function="sum",
        )
        .rename({str(k): v for k, v in {282: "turnover_meur", 284: "turnover_mkr", 285: "sales_mkr", 287: "purchases_mkr"}.items()})
    )
    # series that never traded in a month are absent from the pivot -> fill 0
    for col in ("turnover_meur", "turnover_mkr", "sales_mkr", "purchases_mkr"):
        if col not in monthly.columns:
            monthly = monthly.with_columns(pl.lit(0.0).alias(col))

    monthly = monthly.with_columns(
        (pl.col("purchases_mkr") - pl.col("sales_mkr")).alias("net_purchases_mkr")
    )

    # month-end EUR mid and month-end reserves
    eur_mid = (
        eur.sort("date")
        .group_by("month")
        .agg(pl.col("eur_mid").last().alias("isk_per_eur"))
    )
    reserves = (
        res.sort("date")
        .group_by("month")
        .agg(pl.col("reserves_mkr").last().alias("reserves_mkr"))
    )

    out = (
        monthly.join(eur_mid, on="month", how="left")
        .join(reserves, on="month", how="left")
        .sort("month")
        .with_columns([
            pl.col("month").alias("date"),
            pl.lit("CBI FX intervention: net purchases (kaup-sala), market turnover, reserves, ISK/EUR mid").alias("item_en"),
        ])
        .select([
            "date",
            "net_purchases_mkr",
            "turnover_mkr",
            "reserves_mkr",
            "isk_per_eur",
            "item_en",
        ])
    )
    return out


def cmd_list() -> None:
    with _client() as client:
        path = fetch_fx_market(client)
    df = read_fx_market(path)
    print("\nDaily FX series available (monthly sums used downstream):")
    for row in (
        df.group_by("series_id")
        .agg([pl.col("value_mkr").count().alias("days"), pl.col("date").min().alias("from"), pl.col("date").max().alias("to")])
        .sort("series_id")
        .iter_rows()
    ):
        print(f"  TS {row[0]:<4} {FX_SERIES[row[0]]:<45} {row[1]} days  {row[2]} .. {row[3]}")


def cmd_fetch() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    print("Fetching Seðlabanki FX data...")
    with _client() as client:
        fx_path = fetch_fx_market(client)
        eur_path = fetch_eur_mid(client)
        cb_path = fetch_cb_balance_sheet(client)

    df = build_monthly(fx_path, eur_path, cb_path)
    df.write_csv(OUT_CSV)
    print(f"\nWrote {len(df)} monthly rows to {OUT_CSV}")
    print("\nCOVID era (2020-01 .. 2021-12):")
    covid = df.filter((pl.col("date") >= date(2020, 1, 1)) & (pl.col("date") <= date(2021, 12, 31)))
    with pl.Config(tbl_rows=30, tbl_width_chars=110):
        print(covid)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="cmd", required=True)
    sub.add_parser("list", help="list the daily FX series found upstream")
    sub.add_parser("fetch", help="fetch raw + write data/processed/sedlabanki_fx_intervention.csv")
    args = parser.parse_args()
    if args.cmd == "list":
        cmd_list()
    else:
        cmd_fetch()


if __name__ == "__main__":
    main()
