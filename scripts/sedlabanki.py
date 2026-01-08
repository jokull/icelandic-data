"""
Process Seðlabanki (Central Bank of Iceland) Excel downloads into tidy CSVs.

Wide format (dates as columns) -> Long format (date column + value column)
"""

import re
from pathlib import Path

import polars as pl
from openpyxl import load_workbook


RAW_DIR = Path(__file__).parent.parent / "data" / "raw" / "sedlabanki"
PROCESSED_DIR = Path(__file__).parent.parent / "data" / "processed"


def process_newcredit():
    """Process new credit by sector data."""
    xlsx_path = RAW_DIR / "newcredit.xlsx"
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        return

    print(f"Processing {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)

    # Find the data sheet (usually 'I')
    ws = wb["I"]

    # Find header row with "M.kr." in first column
    header_row = None
    for row_num, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        if row[0] and "M.kr" in str(row[0]):
            header_row = row_num
            break

    if header_row is None:
        print("Could not find header row with M.kr.")
        return

    print(f"Found header at row {header_row}")

    # Build headers from the header row
    headers = ["sector"]
    for cell in list(ws[header_row])[1:]:
        if cell.value is None:
            break
        val = cell.value
        if hasattr(val, "strftime"):
            headers.append(val.strftime("%Y-%m"))
        else:
            headers.append(str(val))

    print(f"Found {len(headers) - 1} date columns")

    # Extract data rows (starting from header_row + 1)
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:
            continue
        sector = str(row[0]).strip()
        if not sector or sector.startswith("*"):
            continue
        row_data = {"sector": sector}
        for i, val in enumerate(row[1:], start=1):
            if i < len(headers):
                row_data[headers[i]] = val
        rows.append(row_data)

    if not rows:
        print("No data rows found")
        return

    # Create wide DataFrame
    df_wide = pl.DataFrame(rows)

    # Get date columns (exclude 'sector')
    date_cols = [c for c in df_wide.columns if c != "sector" and re.match(r"\d{4}-\d{2}", c)]

    # Unpivot to long format
    df_long = df_wide.unpivot(
        index="sector",
        on=date_cols,
        variable_name="date",
        value_name="value_mkr",
    )

    # Parse date and clean up (handle empty strings as nulls)
    df_long = df_long.with_columns([
        pl.col("date").str.to_date("%Y-%m").alias("date"),
        pl.when(pl.col("value_mkr") == "")
        .then(None)
        .otherwise(pl.col("value_mkr"))
        .cast(pl.Float64)
        .alias("value_mkr"),
    ])

    # Extract English sector name if available (after " / ")
    df_long = df_long.with_columns([
        pl.when(pl.col("sector").str.contains(" / "))
        .then(pl.col("sector").str.split(" / ").list.last())
        .otherwise(pl.col("sector"))
        .alias("sector_en"),
        pl.when(pl.col("sector").str.contains(" / "))
        .then(pl.col("sector").str.split(" / ").list.first())
        .otherwise(pl.col("sector"))
        .alias("sector_is"),
    ])

    # Sort by date and sector
    df_long = df_long.sort(["date", "sector"])

    # Save
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out_path = PROCESSED_DIR / "sedlabanki_newcredit.csv"
    df_long.write_csv(out_path)
    print(f"Wrote {len(df_long)} rows to {out_path}")

    # Preview
    print("\nSample:")
    print(df_long.head(10))


def process_balance_sheets():
    """Process deposit institution balance sheet data."""
    xlsx_path = RAW_DIR / "balance_sheets.xlsx"
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        return

    print(f"Processing {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)

    ws = wb["INN_BALANCE_SHEETS_TOTAL"]

    # Row 3 has dates starting from column C
    # Row 4+ has data with labels in column B
    dates = []
    for cell in list(ws[3])[2:]:  # Skip columns A, B
        if cell.value is None:
            break
        dates.append(str(cell.value))

    print(f"Found {len(dates)} date columns ({dates[0]} to {dates[-1]})")

    # Extract data rows (skip footer notes)
    skip_prefixes = ("Nýjustu", "Ekki er", "Heimild", "*")
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        label = row[1]  # Column B
        if label is None:
            continue

        label = str(label).strip()
        if not label:
            continue

        # Skip footer notes
        if any(label.startswith(p) for p in skip_prefixes):
            continue

        # Calculate hierarchy level from leading spaces
        stripped = label.lstrip()
        level = (len(label) - len(stripped)) // 2

        row_data = {"item": stripped, "level": level}
        for i, val in enumerate(row[2:]):  # Skip columns A, B
            if i < len(dates):
                row_data[dates[i]] = val
        rows.append(row_data)

    if not rows:
        print("No data rows found")
        return

    print(f"Found {len(rows)} balance sheet items")

    # Create wide DataFrame
    df_wide = pl.DataFrame(rows)

    # Unpivot to long format
    df_long = df_wide.unpivot(
        index=["item", "level"],
        on=dates,
        variable_name="date",
        value_name="value_mkr",
    )

    # Parse date (format: YYYY-MM) and handle non-numeric values
    df_long = df_long.with_columns([
        pl.col("date").str.to_date("%Y-%m").alias("date"),
        pl.col("value_mkr").cast(pl.Float64, strict=False),
    ])

    # Extract English name if available (after " / ")
    df_long = df_long.with_columns([
        pl.when(pl.col("item").str.contains(" / "))
        .then(pl.col("item").str.split(" / ").list.last())
        .otherwise(pl.col("item"))
        .alias("item_en"),
        pl.when(pl.col("item").str.contains(" / "))
        .then(pl.col("item").str.split(" / ").list.first())
        .otherwise(pl.col("item"))
        .alias("item_is"),
    ])

    # Sort by date and item
    df_long = df_long.sort(["date", "item"])

    # Save
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out_path = PROCESSED_DIR / "sedlabanki_balance_sheets.csv"
    df_long.write_csv(out_path)
    print(f"Wrote {len(df_long)} rows to {out_path}")

    # Preview
    print("\nSample (Assets total):")
    sample = df_long.filter(pl.col("item_en") == "Assets, total").tail(12)
    print(sample)


def main():
    """Process all Seðlabanki datasets."""
    print("=" * 60)
    print("Processing Seðlabanki data")
    print("=" * 60)

    process_balance_sheets()
    print()
    process_newcredit()

    print()
    print("Done!")


if __name__ == "__main__":
    main()
